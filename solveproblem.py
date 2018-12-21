import pandas as pd
import sys
import os
import time
import pyodbc
import decimal

from pulp import *
from openpyxl import load_workbook
from itertools import permutations
from tqdm import tqdm

"""
Driver download here:
https://www.microsoft.com/en-us/download/details.aspx?id=54920
To clean installation from 32 bit software:
https://support.microsoft.com/en-us/help/17588/fix-problems-that-block-programs-from-being-installed-or-removed
Careful with current Excel installation
"""


def read_data_from_access(path, table, query=None):
    """
    reads table from provided access db path into pandas dataframe
    """
    string = (r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};'+
              'DBQ=' + path)
    conn = pyodbc.connect(string)
    cursor = conn.cursor()

    if query is None:
        query = f'select * from {table}'  # Q_EstraiPerOrdinamento, Variabili
    cursor.execute(query)

    data = cursor.fetchall()
    conn.close()

    columns = [c[0] for c in cursor.description]
    types = [c[1] for c in cursor.description]

    df = pd.DataFrame([list(row) for row in data], columns=columns)
    for i in range(len(columns)):
        if types[i] in [decimal.Decimal, int]:
            df[columns[i]] = df[columns[i]].astype(float)
    return df


def read_data(file_path, macchina=None, db=False, cost=False, filter_columns=True):
    if db:
        data = read_data_from_access(file_path, 'Q_EstraiPerOrdinamento')
        columns_data = read_data_from_access(file_path, 'VariabiliMacchina')
        if macchina is not None:
            data = data.loc[data['Macchina']==macchina].copy().sort_values('Prog')
            columns_data = columns_data.loc[columns_data['Macchina']==macchina].copy()
            data.drop('Macchina', axis=1, inplace=True)
            columns_data.drop('Macchina', axis=1, inplace=True)
    else:
        file_obj = pd.ExcelFile(file_path)
        columns_data = pd.read_excel(file_obj, 'Variabili')
        
        data = pd.read_excel(file_obj, 'Foglio1')
        
    data.columns = [c.lower() for c in data.columns]
    columns_data['Variabile'] = columns_data['Variabile'].str.lower()
    to_consider = columns_data.loc[
        columns_data['ConsideraInOttimizzazione 1/0']==1, 
        'Variabile'].tolist()
    data = data.set_index('prog')
    data['nrcol'] = data['nrcol'].astype(float)

    for col in data.columns:
        if type(data[col].iloc[0])==str:
            data[col] = data[col].fillna('None')
        else:
            data[col] = data[col].fillna(0)

    if filter_columns:
        data = data[to_consider].copy()

    if cost:
        cost_series = columns_data[
            ['Variabile', 'ValoreCostoCambio', 'Svuota', 'Riempi', 'CambioManica']].set_index('Variabile')
        return cost_series
    else:
        return data


def calculate_switch_costs(data, file_path, db, macchina=None, all_comb=True):
    """calculates cost for switching between any two configurations
    """

    cost_series = read_data(file_path, db=db, macchina=macchina, cost=True)
    switch_costs = {}

    data = data.reset_index()
    if all_comb:
        iterlist = permutations(data.index.tolist(), 2)
    else:
        iterlist = [(data.index[i-1], data.index[i]) 
                    for i in range(1, len(data))]

    for cpc1, cpc2 in iterlist:
        try:
            row1 = data.loc[cpc1]
            row2 = data.loc[cpc2]

            prog1 = row1['prog']
            prog2 = row2['prog']
            row1 = row1.drop('prog')
            row2 = row2.drop('prog')
            row1_copy = row1.fillna(row2)
            row2_copy = row2.fillna(row1)

            switch_cost = ((~(row1_copy==row2_copy))*1*cost_series[
                'ValoreCostoCambio']).sum()
            
            # handle nans
            with_nans = (pd.isnull(row1) | pd.isnull(row2))
            if with_nans.any():
                sel_1 = row1[row1.index[with_nans]]
                sel_2 = row2[row2.index[with_nans]]

                if ((pd.isnull(sel_1).any()) & 
                    (not pd.isnull(sel_2).any())):
                    # riempi 
                    switch_cost += cost_series.loc[row1.index[with_nans], 
                                                'Riempi'].sum()
                elif ((pd.isnull(sel_2).any()) & 
                        (not pd.isnull(sel_1).any())):
                    # svuota 
                    switch_cost += cost_series.loc[row2.index[with_nans], 
                                                'Svuota'].sum()               
        except:
            import ipdb; ipdb.set_trace()    
        switch_costs[(prog1, prog2)] = switch_cost 

    return switch_costs


def create_problem_binary(data, switch_costs, source=123, target=127):
    """Create pulp problem to minimize costs, given source and target
    """

    prob = LpProblem('order_problem', LpMinimize)

    graph_edges = list(itertools.permutations(data.index.tolist(), 2))
    nodes = data.index.tolist()

    if source:
        nodes = [source] + [n for n in nodes if n not in [source, target]] + [target]

    var_dict = {}
    for cpc1, cpc2 in graph_edges:
        x=LpVariable(f"choice_{cpc1}_{cpc2}", 0, 1, LpBinary)
        var_dict[(cpc1, cpc2)] = x
    
    dummy_dict = {node: LpVariable(f"dummy_{node}", None, None, LpInteger) for node in nodes}
    prob += lpSum([switch_costs[couple]*var_dict[couple] for couple in var_dict])

    for i, j in graph_edges:
        # each edge only in one direction
        prob += var_dict[i, j] + var_dict[j, i] <= 1

    for node in nodes:
        if node == source:
            # sum of ingoing = 0, sum of outgoing = 1
            prob += lpSum([var_dict[i, j] for i, j in graph_edges if j==node]) == 0
            prob += lpSum([var_dict[j, i] for i, j in graph_edges if j==node]) == 1
        elif node == target:
            # sum of ingoing = 1, sum of outgoing = 0
            prob += lpSum([var_dict[i, j] for i, j in graph_edges if j==node]) == 1
            prob += lpSum([var_dict[j, i] for i, j in graph_edges if j==node]) == 0
        else:
            # sum of ingoing = 1, sum of outgoing = 1
            prob += lpSum([var_dict[i, j] for i, j in graph_edges if j==node]) == 1
            prob += lpSum([var_dict[j, i] for i, j in graph_edges if j==node]) == 1

    # make sure that there are no closed loops 
    # (https://en.wikipedia.org/wiki/Travelling_salesman_problem#Integer_linear_programming_formulation)
    for i in nodes[1:]:
        prob += dummy_dict[i]<=len(nodes)-1
        prob += dummy_dict[i]>=0        
        for j in nodes[1:]:
            if i!=j:
                prob += dummy_dict[i]-dummy_dict[j] + len(nodes)*var_dict[(i, j)] <= len(nodes)-1
    
    return prob, var_dict


def solve_problem(file_path, max_sec=300, macchina=None, row_from=None, row_until=None, source=None, target=None):
    """solves problem given cost and start-end
    """

    db = ('accdb' in file_path)
    data = read_data(file_path, db=db, macchina=macchina)

    if source is None:
        data = data.loc[row_from:row_until].copy()
    print(f'Sorting {len(data)} rows')
    
    switch_costs = calculate_switch_costs(data, file_path, db=db, macchina=macchina)
    prob, var_dict = create_problem_binary(data, switch_costs, source=source, target=target)
    start = time.time()
    prob.solve(PULP_CBC_CMD(maxSeconds=max_sec, 
                            fracGap = 0.01))  # takes kwargs, check solve_CBC
    
    print("Status:", LpStatus[prob.status])
    print("Elapsed", round(time.time()-start, 2), "seconds")

    for file in os.listdir():
        if (('.mps' in file) or ('.sol' in file)):
            os.remove(file)
    
    if prob.status==1:
        sol = pd.DataFrame(
            [[couple[0], couple[1], var_dict[couple].value()] for couple in var_dict],
            columns=['da', 'a', 'value'])
        sol = sol.loc[sol.value>0]
        
        if source is None:
            source = data.index[0]
        order = [source]

        while len(order)<=len(sol):
            next_value = sol.loc[sol.da==order[-1], 'a'].values[0]
            order.append(next_value)

        print('Costo minimo:', prob.objective.value())

        export_result(order, file_path, db, switch_costs, macchina=macchina)
        return True


def export_result(order, file_path, db, switch_costs, macchina=None):
    """Exports result to excel
    """

    data = read_data(file_path, db=db, filter_columns=False, macchina=macchina)
    cost_series = read_data(file_path, db=db, cost=True, macchina=macchina)

    to_change = data.loc[pd.Index(order)].copy()
    remain_same = data.drop(pd.Index(order), axis=0)
    to_change = to_change.reset_index().reset_index()
    to_change['index'] = to_change['index']+order[0]
    new_data = remain_same.reset_index().append(to_change, sort=False)

    # check if new index contains also old numbers which would cause duplicates
    new_data = add_cost(new_data, cost_series, switch_costs)
    import ipdb; ipdb.set_trace()
    new_data['index'] = new_data['index'].fillna(new_data['prog']).astype(int)    
    
    if db:
        new_data = new_data.rename(columns={'index': 'new_prog'})
        new_data = new_data.rename(columns={'prog': 'old_prog'})
        import ipdb; ipdb.set_trace()
        write_order_to_db(file_path, new_data, macchina)
    else:
        new_data.drop('prog', axis=1, inplace=True)
        new_data.rename(columns={'index': 'prog'}, inplace=True)
        new_data = new_data[
            ['prog'] + 
            [col for col in new_data.columns if col not in ['prog']]]
        new_data.so
        save_file(file_path, new_data)


def write_order_to_db(file_path, new_data, macchina):
    
    changed = new_data.loc[new_data['new_prog']!=new_data['old_prog']].copy()
    string = (r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};'+
              'DBQ=' + file_path)
    conn = pyodbc.connect(string)
    cursor = conn.cursor()
    import ipdb; ipdb.set_trace()
    for idx, row in tqdm(changed.iterrows()):

        # query_select = f"""select * from Q_EstraiPerOrdinamento
        #                 where 
        #                     (
        #                         Macchina='{macchina}' and 
        #                         Prog={row['old_prog']} and 
        #                         Var_K='{row['var_k']}' and 
        #                         NumeroCom={row['numerocom']} and 
        #                         IDRIGA={row['idriga']} and 
        #                         Fascia='{row['fascia']}' and 
        #                         NrCol='{row['nrcol']}' and 
        #                         QTS='{row['qts']}' and
        #                         S='{row['s']}' and
        #                         T='{row['t']}' and
        #                         DescCol01='{row['desccol01']}'
        #                     )"""
        # sel = read_data_from_access(file_path, 'Q_EstraiPerOrdinamento', query_select)
        
        query = f"""UPDATE Q_EstraiPerOrdinamento
            SET Prog={row['new_prog']} where 
            (
                Macchina='{macchina}' and 
                Prog={row['old_prog']} and 
                Var_K='{row['var_k']}' and 
                NumeroCom={row['numerocom']} and 
                IDRIGA={row['idriga']} and 
                Fascia='{row['fascia']}' and 
                NrCol='{row['nrcol']}' and 
                QTS='{row['qts']}' and
                S='{row['s']}' and
                T='{row['t']}' and
                DescCol01='{row['desccol01']}'
            )"""
        cursor.execute(query)
    conn.commit()
    conn.close()


def write_cost_to_db(file_path, macchina, data):
    string = (r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};'+
              'DBQ=' + file_path)
    conn = pyodbc.connect(string)
    cursor = conn.cursor()

    data['_c_for'] = data['_c_for'].fillna(0)
    for prog in tqdm(range(len(data))):
        query = (f"UPDATE Q_EstraiPerOrdinamento " + 
                 f"SET _C_for={data.iloc[prog]['_c_for']} where (" + 
                 f"Macchina='{macchina}' and " +
                 f"Prog={data.index[prog]})")
        cursor.execute(query)        
    conn.commit()
    conn.close()


def save_file(file_path, new_data, db=False):
    if db:
        pass
    else:
        book = load_workbook(file_path)
        if 'results' in book.sheetnames:
            book.remove(book['results'])
            book.save(file_path)
            book = load_workbook(file_path)
        
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book

        new_data.to_excel(writer, 'results', index=False)
        writer.save()


def add_cost(new_data, cost_series, switch_costs):
    cost = cost_series.loc['nrcol', ['ValoreCostoCambio', 'CambioManica']].sum()

    for i in tqdm(range(1, len(new_data))):
        idx = new_data.index[i]
        idx_prev = new_data.index[i-1]
        if (idx_prev, idx) in switch_costs:
            new_data.loc[idx, '_c_for'] = switch_costs[(idx_prev, idx)]

    new_data['_c_for'] = new_data['_c_for'] + new_data['nrcol'].astype(float)*cost 
    return new_data


def calculate_cost_unsorted(file_path, macchina):
    db = ('accdb' in file_path)
    data = read_data(file_path, db=db, filter_columns=False, macchina=macchina)
    cost_series = read_data(file_path, db=db, cost=True, macchina=macchina)
    switch_costs = calculate_switch_costs(data, file_path, db, macchina=macchina, all_comb=False)
    data = add_cost(data, cost_series, switch_costs)

    if db:
        write_cost_to_db(file_path, macchina, data)
    else:
        save_file(file_path, data)


if __name__=='__main__':
    
    if getattr(sys, 'frozen', False):
        # frozen
        dir_ = os.path.dirname(sys.executable)
    else:
        # unfrozen
        dir_ = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(dir_, 'DbOrdinamento.accdb')
    print(file_path)

    if len(sys.argv)>1:
        arguments = sys.argv[1:]
        macchina = arguments[0]
        task = arguments[1]  # "order" or "cost"
    else:
        macchina = 'Diamond'
        task = 'order'
        arguments = [macchina, task, 500]
    
    if task=='order':      
        max_sec = int(arguments[2])      
        row_from = int(input('Specificare ordinamento di inizio: '))
        row_until = int(input('Specificare ordinamento di fine (escluso): '))
        order = solve_problem(
            file_path, macchina=macchina, 
            row_from=row_from, row_until=row_until, max_sec=max_sec)
    elif task=='cost':
        calculate_cost_unsorted(file_path, macchina)

